#!/usr/bin/env python3
# ============================================================
# gerar_quizizz.py — Gerador de planilha Quizizz Paper Mode
# Prof. Hérisson Chaves — Ciências da Natureza
#
# Implementa as regras técnicas da skill global `quizizz-paper`
# COM o override pedagógico desta skill local: todas as questões
# no nível Médio (sem prefixos [FÁCIL]/[MÉDIA]/[DIFÍCIL]),
# linhas em amarelo claro (com laranja para questões mais
# desafiadoras).
#
# USO:
#   python3 gerar_quizizz.py \
#     --tema "Pressão Hidrostática" \
#     --data 20250615 \
#     --questoes saida/questoes_temp.json \
#     --saida saida/
#
#   # ou em modo demo (sem JSON):
#   python3 gerar_quizizz.py --demo --tema "Demo" --data 20250615 --saida saida/
# ============================================================

import argparse
import json
import os
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
except ImportError:
    print("❌ openpyxl não instalado. Rode: pip install openpyxl --break-system-packages")
    sys.exit(1)


HEADER_FILL = "4472C4"      # azul header
ROW_MEDIO = "FFF3C4"        # amarelo claro (Médio — padrão desta skill)
ROW_DIFICIL = "FFE0B2"      # laranja claro (excepcional, p/ questões mais desafiadoras)

LIMITE_ENUNCIADO = 500
LIMITE_ALTERNATIVA = 200


def slugify(texto: str) -> str:
    """Converte 'Pressão Hidrostática' em 'pressao_hidrostatica'."""
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r"[^A-Za-z0-9]+", "_", texto).strip("_").lower()
    return texto or "fisica"


def truncar(texto: str, limite: int, label: str = "campo") -> tuple[str, bool]:
    """Retorna (texto_resumido, foi_truncado)."""
    texto = (texto or "").strip()
    if len(texto) <= limite:
        return texto, False
    cortado = texto[: limite - 3].rsplit(" ", 1)[0] + "..."
    return cortado, True


def validar_questao(q: dict, idx: int) -> list[str]:
    """Retorna lista de problemas (vazia se OK)."""
    problemas = []
    if not q.get("enunciado"):
        problemas.append(f"Q{idx}: enunciado vazio")
    opcoes = q.get("opcoes") or []
    if len(opcoes) != 4:
        problemas.append(f"Q{idx}: tem {len(opcoes)} alternativas (precisa de exatamente 4)")
    correta = q.get("correta")
    if not isinstance(correta, int) or correta < 1 or correta > 4:
        problemas.append(f"Q{idx}: correta inválida ({correta!r}); precisa ser 1..4")
    return problemas


def carregar_questoes(caminho_json: str) -> list[dict]:
    if not os.path.exists(caminho_json):
        raise FileNotFoundError(f"JSON de questões não encontrado: {caminho_json}")
    with open(caminho_json, "r", encoding="utf-8") as f:
        return json.load(f)


def questoes_demo() -> list[dict]:
    """Conjunto de demonstração — usado se o JSON não existir."""
    return [
        {
            "enunciado": "Um mergulhador desce de 2 m para 4 m. Como a pressão hidrostática sobre ele varia?",
            "opcoes": [
                "Dobra, porque a profundidade dobrou",
                "Aumenta, mas a relação não é direta",
                "Permanece igual — não depende da profundidade",
                "Reduz à metade — a coluna acima 'empurra' menos",
            ],
            "correta": 1,
            "conceito": "Pressão Hidrostática",
        },
        {
            "enunciado": "Dois recipientes têm formatos diferentes mas a mesma altura de coluna de água. Em qual a pressão no fundo é maior?",
            "opcoes": [
                "No mais largo, porque tem mais volume",
                "No mais estreito, porque a água 'aperta' mais",
                "Iguais — só a altura da coluna importa",
                "Depende do material do recipiente",
            ],
            "correta": 3,
            "conceito": "Paradoxo Hidrostático",
        },
        {
            "enunciado": "Um macaco hidráulico tem o pistão pequeno com 1/10 da área do grande. Aplicando 200 N no pequeno, a força no grande é:",
            "opcoes": [
                "20 N — força é dividida pela razão das áreas",
                "200 N — pressão se conserva, força também",
                "2000 N — pressão se conserva, força multiplicada pela razão",
                "Não dá para saber sem a altura",
            ],
            "correta": 3,
            "conceito": "Princípio de Pascal",
        },
        {
            "enunciado": "A base de uma represa é mais espessa que o topo. Por quê?",
            "opcoes": [
                "Para sustentar o peso da estrutura",
                "Porque a pressão hidrostática é maior no fundo",
                "Para evitar a evaporação",
                "Por questões estéticas — não há razão física",
            ],
            "correta": 2,
            "conceito": "Aplicação de P = ρgh",
        },
    ]


def gerar_xlsx(questoes: list[dict], tema: str, data: str, pasta_saida: str) -> tuple[str, list[str]]:
    """Gera o .xlsx Quizizz e retorna (caminho_arquivo, lista_avisos)."""

    avisos = []
    # Validar
    for i, q in enumerate(questoes, 1):
        for p in validar_questao(q, i):
            avisos.append(f"⚠️  {p}")
    if any("inválida" in a or "alternativas" in a or "vazio" in a for a in avisos):
        # Erros bloqueantes
        print("Erros nas questões:")
        for a in avisos:
            print("  " + a)
        raise ValueError("Questões inválidas — aborto.")

    # Truncar enunciados/alternativas longos
    for i, q in enumerate(questoes, 1):
        enun, cortado = truncar(q["enunciado"], LIMITE_ENUNCIADO, f"Q{i} enunciado")
        if cortado:
            avisos.append(f"⚠️  Q{i}: enunciado adaptado de {len(q['enunciado'])} para {len(enun)} chars")
        q["enunciado"] = enun
        for j, opc in enumerate(q["opcoes"]):
            opc2, cort = truncar(opc, LIMITE_ALTERNATIVA, f"Q{i} opção {j+1}")
            if cort:
                avisos.append(f"⚠️  Q{i} opção {j+1}: adaptada de {len(opc)} para {len(opc2)} chars")
            q["opcoes"][j] = opc2

    # Construir planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Create a Quiz"

    headers = [
        "Question Text", "Question Type",
        "Option 1", "Option 2", "Option 3", "Option 4", "Option 5",
        "Correct Answer", "Time in seconds", "Image Link", "Answer explanation",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True, color="FFFFFF")
    border_thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # Larguras
    larguras = [60, 18, 30, 30, 30, 30, 10, 16, 18, 24, 24]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Dados — todas as linhas em ROW_MEDIO por padrão (override do override)
    for q in questoes:
        # Override desta skill: NUNCA prefixar com [MÉDIA] / [FÁCIL] / [DIFÍCIL]
        # Enunciado vai limpo para o aluno
        linha = [
            q["enunciado"],
            "Multiple Choice",
            q["opcoes"][0], q["opcoes"][1], q["opcoes"][2], q["opcoes"][3],
            "",                         # Option 5 sempre vazio
            int(q["correta"]),
            30,                         # Time in seconds fixo
            "",                         # Image Link sempre vazio
            "",                         # Answer explanation sempre vazio
        ]
        ws.append(linha)
        # Cor de fundo da linha
        cor = ROW_DIFICIL if q.get("desafiadora") else ROW_MEDIO
        fill = PatternFill("solid", fgColor=cor)
        for col in range(1, 12):
            cell = ws.cell(ws.max_row, col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = fill
            cell.border = border_thin

    # Altura mínima das linhas de dados
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 60

    # Salvar
    Path(pasta_saida).mkdir(parents=True, exist_ok=True)
    nome = f"quizizz_{slugify(tema)}_{data}.xlsx"
    caminho = os.path.join(pasta_saida, nome)
    wb.save(caminho)
    return caminho, avisos


def imprimir_gabarito(questoes: list[dict]) -> None:
    print("\n📋 Gabarito Quizizz:")
    letras = ["A", "B", "C", "D"]
    for i, q in enumerate(questoes, 1):
        idx = q["correta"] - 1
        opc = q["opcoes"][idx]
        opc_curto = opc if len(opc) <= 50 else opc[:47] + "..."
        conceito = q.get("conceito", "—")
        print(f"  Q{i} [{conceito}] → {letras[idx]}) {opc_curto}")


def main():
    p = argparse.ArgumentParser(description="Gera planilha .xlsx Quizizz Paper Mode (override Médio).")
    p.add_argument("--tema", required=True, help="Tema da aula (vai para o nome do arquivo)")
    p.add_argument("--data", required=True, help="Data no formato YYYYMMDD")
    p.add_argument("--questoes", help="Arquivo JSON com a lista de questões")
    p.add_argument("--demo", action="store_true", help="Usar conjunto de demonstração (4 questões)")
    p.add_argument("--saida", default="saida", help="Pasta de saída (default: saida)")
    args = p.parse_args()

    if args.demo:
        questoes = questoes_demo()
        print("ℹ️  Modo demo — usando 4 questões de demonstração de Hidrostática.")
    elif args.questoes:
        questoes = carregar_questoes(args.questoes)
    else:
        print("❌ Forneça --questoes <arquivo.json> ou --demo")
        sys.exit(1)

    if not questoes:
        print("❌ Nenhuma questão encontrada.")
        sys.exit(1)

    caminho, avisos = gerar_xlsx(questoes, args.tema, args.data, args.saida)
    print(f"\n✅ Planilha gerada: {caminho} ({len(questoes)} questões)")
    if avisos:
        print("\nAvisos:")
        for a in avisos:
            print("  " + a)

    imprimir_gabarito(questoes)
    print("\n💡 Importe no Quizizz: Criar → Importar de planilha → Paper Mode")


if __name__ == "__main__":
    main()
